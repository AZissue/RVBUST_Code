import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

class DataManager:
    def __init__(self):
        self.current_excel_path = self.get_default_excel_path()
    
    def get_default_excel_path(self):
        """获取默认Excel文件路径"""
        # 使用C盘用户文件夹作为存储位置
        user_dir = os.path.expanduser("~")
        data_dir = os.path.join(user_dir, "WorkLog Pro", "data")
        return os.path.join(data_dir, "worklog.xlsx")
    
    def get_excel_file_path(self):
        """获取Excel文件路径"""
        return self.current_excel_path
    
    def set_excel_file_path(self, path):
        """设置Excel文件路径"""
        self.current_excel_path = path
    
    def load_data_from_excel(self):
        """从Excel文件加载数据"""
        file_path = self.get_excel_file_path()
        
        try:
            if os.path.exists(file_path):
                df = pd.read_excel(file_path)
                # 将NaN值替换为空字符串
                df = df.fillna('')
                return df
            else:
                return None
        except Exception as e:
            raise Exception(f"加载Excel文件失败: {str(e)}")
    
    def save_data_to_excel(self, data):
        """保存数据到Excel文件"""
        file_path = self.get_excel_file_path()
        
        try:
            # 确保目录存在
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            df = pd.DataFrame(data)
            # 使用openpyxl引擎保存，以便后续设置格式
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # 获取worksheet对象
                worksheet = writer.sheets['Sheet1']
                
                # 设置列宽和自动换行
                for col_idx, column in enumerate(df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    
                    # ID栏缩短
                    if column == 'ID':
                        worksheet.column_dimensions[col_letter].width = 8
                    # 日期栏缩短
                    elif column == '日期':
                        worksheet.column_dimensions[col_letter].width = 15
                    # 用户名称栏缩短
                    elif column == '用户名称':
                        worksheet.column_dimensions[col_letter].width = 15
                    # 相机型号栏缩短
                    elif column == '相机型号':
                        worksheet.column_dimensions[col_letter].width = 18
                    # 问题类型栏缩短
                    elif column == '问题类型':
                        worksheet.column_dimensions[col_letter].width = 12
                    # 问题进度栏缩短
                    elif column == '问题进度':
                        worksheet.column_dimensions[col_letter].width = 12
                    # 客户问题和解决方法扩大
                    elif column == '客户问题' or column == '解决方法':
                        worksheet.column_dimensions[col_letter].width = 50
                    # 其他列设置合适的宽度
                    else:
                        worksheet.column_dimensions[col_letter].width = 15
                
                # 创建边框样式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 为所有单元格添加自动换行和边框
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                              min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = thin_border
            
            return True
        except Exception as e:
            raise Exception(f"保存Excel文件失败: {str(e)}")
    
    def import_data_from_excel(self, file_path):
        """从指定Excel文件导入数据"""
        try:
            df = pd.read_excel(file_path)
            # 将NaN值替换为空字符串
            df = df.fillna('')
            self.set_excel_file_path(file_path)
            return df
        except Exception as e:
            raise Exception(f"导入Excel文件失败: {str(e)}")
    
    def export_data_to_excel(self, data, file_path):
        """导出数据到指定Excel文件"""
        try:
            # 确保目录存在
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            df = pd.DataFrame(data)
            # 使用openpyxl引擎保存，以便后续设置格式
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # 获取worksheet对象
                worksheet = writer.sheets['Sheet1']
                
                # 设置列宽和自动换行
                for col_idx, column in enumerate(df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    
                    # ID栏缩短
                    if column == 'ID':
                        worksheet.column_dimensions[col_letter].width = 8
                    # 日期栏缩短
                    elif column == '日期':
                        worksheet.column_dimensions[col_letter].width = 15
                    # 用户名称栏缩短
                    elif column == '用户名称':
                        worksheet.column_dimensions[col_letter].width = 15
                    # 相机型号栏缩短
                    elif column == '相机型号':
                        worksheet.column_dimensions[col_letter].width = 18
                    # 问题类型栏缩短
                    elif column == '问题类型':
                        worksheet.column_dimensions[col_letter].width = 12
                    # 问题进度栏缩短
                    elif column == '问题进度':
                        worksheet.column_dimensions[col_letter].width = 12
                    # 客户问题和解决方法扩大
                    elif column == '客户问题' or column == '解决方法':
                        worksheet.column_dimensions[col_letter].width = 50
                    # 其他列设置合适的宽度
                    else:
                        worksheet.column_dimensions[col_letter].width = 15
                
                # 创建边框样式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 为所有单元格添加自动换行和边框
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                              min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = thin_border
            
            return True
        except Exception as e:
            raise Exception(f"导出Excel文件失败: {str(e)}")
